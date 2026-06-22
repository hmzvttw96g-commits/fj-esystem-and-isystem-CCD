#!/usr/bin/env python3
"""功能链④产业转化：智慧芽专利导出 → i_conversion_panel（城市×申请年×C/B/X 件数）。

严格按 config/i_caliber_patent.yml：
  - 计数：按**申请日年份**；发明+实用新型（外观已在检索排除）；同申请号去重计一件。
  - 申请人规则：第一申请人为**企业**计入；**高校/科研院所剔除**（名称正则，含"有限公司"不误杀）；
    个人主口径剔除。信用代码 91=企业纳入、12/11/5x=事业/机关/社团剔除。
  - IPC→C/B/X：任一分类号前缀命中即计入；C⊂B⊂X 嵌套；G06K9(旧模式识别)归 C。
  - 归市：第一申请人(原始申请人)地址→地级市；缺失用工商注册地兜底。
  - 进指数前 per_capita（由 functional_chain 管道除常住人口）。

用法：python3 scripts/build_i_conversion.py --patent-dir ~/Downloads   # 处理 202606*.XLSX
      python3 scripts/build_i_conversion.py --self-test
"""
from __future__ import annotations
import argparse, csv, glob, re, warnings
from collections import defaultdict
from pathlib import Path

warnings.filterwarnings("ignore")
ROOT = Path(__file__).resolve().parents[1]
CFG = ROOT / "config" / "i_caliber_patent.yml"
PANEL = ROOT / "data" / "panel" / "functional_chain" / "i_conversion_panel.csv"
AUDIT = ROOT / "data" / "audit" / "i_conversion"
CITIES = ["福州", "厦门", "泉州", "漳州", "莆田", "三明", "南平", "龙岩", "宁德"]
YEARS = list(range(2019, 2025))
CALIBERS = ["C", "B", "X"]
UNIV_RE = re.compile(r"(大学|学院|研究院|研究所|科学院|科研|职业技术学校)")
COMPANY_RE = re.compile(r"(有限公司|有限责任公司|股份|集团|科技有限|企业)")
EXCLUDE_CODE_PREFIX = ("12", "11", "51", "52", "53")   # 事业单位/机关/社团/民非


def load_ipc_sets():
    import yaml
    q = yaml.safe_load(open(CFG, encoding="utf-8"))
    C = [d["code"] for d in q["C_conservative"]["ipc_prefixes"]] + ["G06K9"]   # 旧码归C
    B = C + [d["code"] for d in q["B_basic"]["additional_ipc_prefixes"]]
    X = B + [d["code"] for d in q["X_expanded"]["additional_ipc_prefixes"]]
    return {"C": C, "B": B, "X": X}


def ipc_caliber(ipc_field, ipc_sets):
    """专利的 IPC 串(如 'G06T3/40 | G06N3/02') → 命中的口径集合(嵌套)。"""
    codes = [c.strip().replace(" ", "") for c in str(ipc_field).split("|") if c.strip()]
    hit = set()
    for cal in CALIBERS:
        if any(any(code.startswith(pre) for pre in ipc_sets[cal]) for code in codes):
            hit.add(cal)
    return hit


def is_enterprise(name, credit_code):
    """第一申请人是否企业（按 config applicant_rule）：高校/院所剔；个人剔；企业纳。"""
    name = str(name or "").strip()
    code = str(credit_code or "").strip()
    if UNIV_RE.search(name) and not COMPANY_RE.search(name):
        return False                                  # 高校/科研院所（白名单:含公司不剔）
    if code[:2] == "91":
        return True                                   # 工商企业
    if code[:2] in EXCLUDE_CODE_PREFIX:
        return False                                  # 事业/机关/社团
    if COMPANY_RE.search(name):
        return True                                   # 名称像企业（无码兜底）
    return False                                      # 个人/不明 → 主口径剔


def attribute_city(*addrs):
    for a in addrs:
        a = str(a or "")
        for c in CITIES:
            if c + "市" in a or ("福建" in a and c in a):
                return c
    return None


def load_patents(patent_dir):
    import openpyxl
    files = sorted(glob.glob(str(Path(patent_dir) / "202606*.XLSX")))
    seen, rows = set(), []
    for p in files:
        wb = openpyxl.load_workbook(p, data_only=True)
        if "专利数据" not in wb.sheetnames:
            wb.close(); continue
        ws = wb["专利数据"]; data = list(ws.iter_rows(values_only=True))
        hdr = [str(c).strip() for c in data[0]]
        for r in data[1:]:
            d = dict(zip(hdr, r))
            ap = str(d.get("申请号", "") or "").strip()
            if not ap or ap in seen:
                continue
            seen.add(ap); rows.append(d)
        wb.close()
    return rows


def run(patent_dir):
    ipc_sets = load_ipc_sets()
    rows = load_patents(patent_dir)
    print(f"去重后专利 {len(rows)} 件。按 config 剔高校/个人、IPC分类、归市…")
    # 计数：cell[(city,year,caliber)] = 件数（嵌套）
    cell = defaultdict(int)
    stat = defaultdict(int)
    for d in rows:
        m = re.match(r"(20\d\d)", str(d.get("申请日", "")))
        if not m:
            stat["无申请日"] += 1; continue
        year = int(m.group(1))
        if year not in YEARS:
            stat["超窗(2025等)"] += 1; continue
        if not is_enterprise(d.get("第一原始申请(专利权)人") or d.get("[标]当前申请(专利权)人"),
                             d.get("工商统一社会信用代码")):
            stat["非企业(高校/个人)剔"] += 1; continue
        cals = ipc_caliber(d.get("IPC分类号", ""), ipc_sets)
        if not cals:
            stat["IPC不入C/B/X"] += 1; continue
        city = attribute_city(d.get("原始申请(专利权)人地址"), d.get("工商注册地址"),
                              d.get("原始申请(专利权)人"))
        if not city:
            stat["归市失败(非9市/地址缺)"] += 1; continue
        stat["入账"] += 1
        for cal in cals:
            cell[(city, year, cal)] += 1
    print("  处理统计:", dict(stat))
    # 写面板
    rows_out = []
    for c in CITIES:
        for y in YEARS:
            for cal in CALIBERS:
                v = cell.get((c, y, cal))
                rows_out.append({"city": c, "year": y, "caliber": cal,
                                 "value": v if v is not None else "",
                                 "population": "", "data_status": "calculated" if v else "calculated_zero"})
    # 保留已有 population 列
    existing = {}
    if PANEL.exists():
        for r in csv.DictReader(open(PANEL, encoding="utf-8-sig")):
            existing[(r["city"], r["year"], r["caliber"])] = r.get("population", "")
    for r in rows_out:
        r["population"] = existing.get((r["city"], str(r["year"]), r["caliber"]), "")
    with open(PANEL, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["city", "year", "caliber", "value", "population", "data_status"])
        w.writeheader(); w.writerows(rows_out)
    print(f"已填 {PANEL.relative_to(ROOT)}（产业转化件数，C/B/X 嵌套）。")
    # 预览 B 口径
    print("\nB口径 专利件数(按申请年):")
    print("市   " + "".join(f"{y:>7}" for y in YEARS))
    for c in CITIES:
        print(f"{c:<4}" + "".join(f"{cell.get((c,y,'B'),0):>7}" for y in YEARS))


def self_test():
    ipc_sets = load_ipc_sets()
    # IPC 分类
    assert ipc_caliber("G06N3/02", ipc_sets) == {"C", "B", "X"}, "G06N→C(嵌套全中)"
    assert ipc_caliber("G06F16/00", ipc_sets) == {"B", "X"}, "G06F16→B(非C)"
    assert ipc_caliber("G06F18/00", ipc_sets) == {"C", "B", "X"}, "G06F18→C"
    assert ipc_caliber("B25J9/00", ipc_sets) == {"X"}, "B25J→仅X"
    assert ipc_caliber("A01B1/00", ipc_sets) == set(), "无关→空"
    assert ipc_caliber("G06K9/00", ipc_sets) == {"C", "B", "X"}, "旧码G06K9→C"
    # 企业识别
    assert is_enterprise("厦门速相科技有限公司", "91350200MA2XRE4H00")
    assert not is_enterprise("华侨大学", "12100000489558586R")
    assert not is_enterprise("厦门大学", "12100000B36952193C")
    assert is_enterprise("某研究院有限公司", "9135...")     # 白名单:含公司不剔
    assert not is_enterprise("张三", "-")                  # 个人剔
    # 归市
    assert attribute_city("362000 福建省泉州市丰泽区") == "泉州"
    assert attribute_city("福建省厦门市思明南路") == "厦门"
    assert attribute_city("北京市海淀区") is None
    print("自测通过：IPC分类(嵌套+旧码)/企业识别(高校剔·白名单)/归市 全部正确。")


def main():
    ap = argparse.ArgumentParser(description="④产业转化 专利→面板")
    ap.add_argument("--patent-dir", default=None)
    ap.add_argument("--self-test", action="store_true")
    args = ap.parse_args()
    if args.self_test:
        self_test()
    elif args.patent_dir:
        run(args.patent_dir)
    else:
        ap.print_help()


if __name__ == "__main__":
    main()
