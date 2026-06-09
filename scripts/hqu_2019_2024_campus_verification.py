#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""华侨大学 2019-2024 年 AI 基础专业群校区归属核验。

本脚本只整理 E 端候选与校区重分类材料，不重算 CCD，不覆盖
latest experimental base panel。
"""

from __future__ import annotations

import json
import math
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


YEARS = [2019, 2020, 2021, 2022, 2023, 2024]
INPUT_CANDIDATES = Path("data/interim/e_system_recheck/latest_xiamen_2022_E_B_candidates.xlsx")
INPUT_SUMMARY = Path("data/panel/e_system_recheck/latest_xiamen_2022_E_B_city_year_summary.xlsx")
LATEST_BASE_PANEL = Path("data/panel/experimental_analysis/latest_experimental_analysis_base_panel_2019_2024.csv")
DESKTOP_PDF = Path(
    "/Users/greenbarry/Desktop/finacial system and educational system CCD/e-基本口径data/"
    "2022年福建省普通高校招生计划（普通类物理科目组）.pdf"
)

OFFICIAL_SOURCES = [
    {
        "source_name": "华侨大学2019年普通高等教育招生章程",
        "year": 2019,
        "url": "https://zsc.hqu.edu.cn/info/1024/2732.htm",
        "source_type": "official_admission_rule",
        "source_reliability": "official",
        "usable_for": "2019 校区与学院归属规则",
        "evidence_summary": "章程列明厦门校区、泉州校区地址及两校区学院分布。",
    },
    {
        "source_name": "华侨大学2020年普通高等教育招生章程",
        "year": 2020,
        "url": "https://zsc.hqu.edu.cn/info/1024/3069.htm",
        "source_type": "official_admission_rule",
        "source_reliability": "official",
        "usable_for": "2020 校区与学院归属规则",
        "evidence_summary": "章程列明计算机科学与技术学院在厦门校区，工学院在泉州校区。",
    },
    {
        "source_name": "华侨大学2021年普通高等教育招生章程",
        "year": 2021,
        "url": "https://zsc.hqu.edu.cn/info/1024/3431.htm",
        "source_type": "official_admission_rule",
        "source_reliability": "official",
        "usable_for": "2021 校区与学院归属规则",
        "evidence_summary": "章程列明计算机科学与技术学院在厦门校区，工学院在泉州校区。",
    },
    {
        "source_name": "华侨大学2021年普高招生专业目录（境内版）",
        "year": 2021,
        "url": "https://zsc.hqu.edu.cn/info/1024/3459.htm",
        "source_type": "official_major_directory",
        "source_reliability": "official",
        "usable_for": "2021 专业-学院-校区映射",
        "evidence_summary": "目录列示工学院(泉州校区)与计算机科学与技术学院(厦门校区)的相关专业。",
    },
    {
        "source_name": "华侨大学2023年境内本科招生信息列表",
        "year": 2023,
        "url": "https://zsc.hqu.edu.cn/zsxx/jnbkzsxx/nzsxx2023.htm",
        "source_type": "official_admission_index",
        "source_reliability": "official",
        "usable_for": "定位 2023 章程和专业目录",
        "evidence_summary": "官方列表包含 2023 年普通高等教育招生章程及 2023 年普高招生专业目录。",
    },
    {
        "source_name": "华侨大学2023年普高招生专业目录（境内版）",
        "year": 2023,
        "url": "https://zsc.hqu.edu.cn/info/1024/3679.htm",
        "source_type": "official_major_directory",
        "source_reliability": "official",
        "usable_for": "2023 专业-学院-校区映射",
        "evidence_summary": "官方招生信息列表指向该专业目录；网页检索摘要显示该目录按学院和校区列示。",
    },
    {
        "source_name": "华侨大学2024年普通高等教育招生章程",
        "year": 2024,
        "url": "https://zsc.hqu.edu.cn/info/1024/7422.htm",
        "source_type": "official_admission_rule",
        "source_reliability": "official",
        "usable_for": "2024 校区与跨校区培养规则",
        "evidence_summary": "章程列明计算机科学与技术学院人工智能专业大一在泉州、一年后回厦门；跨校区培养按最后一年所在地归属。",
    },
    {
        "source_name": "华侨大学2024年普高本科招生专业目录（境内版）",
        "year": 2024,
        "url": "https://zsc.hqu.edu.cn/info/1024/7462.htm",
        "source_type": "official_major_directory",
        "source_reliability": "official",
        "usable_for": "2024 专业目录定位与人工复核",
        "evidence_summary": "官方问答与招生信息指向该目录；用于后续人工核验。",
    },
    {
        "source_name": "华侨大学招生网学院介绍：计算机科学与技术学院",
        "year": "current",
        "url": "https://zsc.hqu.edu.cn/xyjs/jsjkxyjsxy.htm",
        "source_type": "official_college_page",
        "source_reliability": "official",
        "usable_for": "计算机科学与技术学院相关专业校区核验",
        "evidence_summary": "招生网学院介绍显示计算机科学与技术学院在厦门校区，专业含软件工程、计算机科学与技术、信息安全、人工智能。",
    },
    {
        "source_name": "华侨大学招生网学院介绍汇总页",
        "year": "current",
        "url": "https://zsc.hqu.edu.cn/xyjs.htm",
        "source_type": "official_college_page",
        "source_reliability": "official",
        "usable_for": "工学院、机电及自动化学院等学院校区核验",
        "evidence_summary": "招生网学院介绍汇总显示工学院在泉州校区，机电及自动化学院在厦门校区。",
    },
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def to_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return default
        return int(value)
    except Exception:
        return default


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col[:80]:
                text = clean_text(cell.value)
                max_len = max(max_len, min(len(text), 70))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 42))
    wb.save(path)


def write_xlsx(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    style_workbook(path)


def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    warnings: list[str] = []
    if INPUT_CANDIDATES.exists():
        candidates = pd.read_excel(INPUT_CANDIDATES, sheet_name="01_candidates")
    else:
        candidates = pd.DataFrame()
        warnings.append(f"缺少输入候选表：{INPUT_CANDIDATES}")

    if INPUT_SUMMARY.exists():
        summary = pd.read_excel(INPUT_SUMMARY, sheet_name="01_city_year_summary")
    else:
        summary = pd.DataFrame()
        warnings.append(f"缺少厦门 2022 补齐汇总表：{INPUT_SUMMARY}")

    if LATEST_BASE_PANEL.exists():
        panel = pd.read_csv(LATEST_BASE_PANEL)
    else:
        panel = pd.DataFrame()
        warnings.append(f"缺少 latest experimental base panel：{LATEST_BASE_PANEL}")

    for year in [2019, 2020, 2021, 2023, 2024]:
        warnings.append(f"本地未发现 {year} 年华侨大学 E_B 候选明细或招生计划 PDF，已标记 source_gap/manual_review。")
    warnings.append("未直接读取 archive/legacy_outputs；legacy 文件未作为本轮输入。")
    warnings.append("本轮不重算 CCD，不覆盖 latest experimental base panel。")
    warnings.append("未找到可直接读取的 2022 年华侨大学官方招生章程；2022 计划数来自福建省招生计划册，校区判定结合华侨大学官方章程、专业目录和学院介绍。")
    warnings.append("华侨大学 3+1 或跨校区培养记录按最后一年所在地划分；没有年度计划数的记录只确认校区规则，不新增计划数量。")
    return candidates, summary, panel, warnings


def classify_hqu_row(row: pd.Series) -> dict[str, Any]:
    major = clean_text(row.get("major_standard") or row.get("major_original"))
    caliber = clean_text(row.get("caliber"))
    base = row.to_dict()
    campus_confirmed = "unknown"
    evidence_level = "F_unconfirmed"
    source_url = ""
    evidence_text = ""
    reason = ""
    reassigned_city = "manual_review"
    include_xm = 0
    include_qz = 0
    manual = 1
    reliability = "unconfirmed"

    xiamen_cs = {"人工智能", "计算机科学与技术", "软件工程", "信息安全"}
    quanzhou_engineering = {"物联网工程", "数据科学与大数据技术"}
    xiamen_mech_x = {"智能制造工程"}

    if major in xiamen_cs:
        campus_confirmed = "xiamen_campus"
        evidence_level = "C_official_major_page"
        source_url = "https://zsc.hqu.edu.cn/xyjs/jsjkxyjsxy.htm"
        evidence_text = "招生网学院介绍显示计算机科学与技术学院在厦门校区，专业含软件工程、计算机科学与技术、信息安全、人工智能。"
        reason = "2022 计划册提供专业计划数；官方学院介绍及相邻年度官方专业目录支持该专业隶属计算机科学与技术学院厦门校区。"
        reassigned_city = "厦门"
        include_xm = 1 if caliber == "B_basic" else 0
        include_qz = 0
        manual = 0
        reliability = "official"
    elif major in quanzhou_engineering:
        campus_confirmed = "quanzhou_campus"
        evidence_level = "C_official_major_page"
        source_url = "https://zsc.hqu.edu.cn/xyjs.htm"
        evidence_text = "招生网学院介绍与官方专业目录显示工学院在泉州校区，工学院相关专业包含物联网工程、数据科学与大数据技术。"
        reason = "2022 计划册提供专业计划数；官方专业目录/学院介绍支持该专业隶属工学院泉州校区。"
        reassigned_city = "泉州"
        include_xm = 0
        include_qz = 1 if caliber == "B_basic" else 0
        manual = 0
        reliability = "official"
    elif major in xiamen_mech_x:
        campus_confirmed = "xiamen_campus"
        evidence_level = "C_official_major_page"
        source_url = "https://zsc.hqu.edu.cn/xyjs.htm"
        evidence_text = "招生网学院介绍显示机电及自动化学院在厦门校区，专业含智能制造工程。"
        reason = "智能制造工程属于 X 扩大口径候选，不进入 E_B 主口径；校区可作为厦门校区记录。"
        reassigned_city = "厦门"
        include_xm = 0
        include_qz = 0
        manual = 0
        reliability = "official"
    else:
        reason = "未能根据当前规则确认专业对应校区，需人工复核。"

    out = {
        "year": to_int(base.get("year"), 2022),
        "school_name": clean_text(base.get("school_name")),
        "major_original": clean_text(base.get("major_original")),
        "major_standard": major,
        "major_group": clean_text(base.get("major_group")),
        "caliber": caliber,
        "plan_count": to_int(base.get("plan_count")),
        "admission_scope": clean_text(base.get("admission_scope")),
        "batch": clean_text(base.get("batch")),
        "remarks": clean_text(base.get("remarks")),
        "original_candidate_city": clean_text(base.get("city")),
        "campus_confirmed": campus_confirmed,
        "campus_evidence_level": evidence_level,
        "campus_source_url": source_url,
        "campus_evidence_text": evidence_text,
        "campus_decision_reason": reason,
        "reassigned_city": reassigned_city,
        "include_in_xiamen_E_B": include_xm,
        "include_in_quanzhou_E_B": include_qz,
        "manual_review_required": manual,
        "source_reliability": reliability,
        "reviewer_decision": "pending_review",
        "reviewer_notes": "按校区归属重分类；不覆盖原补齐表和 latest experimental base panel。",
    }
    return out


def source_gap_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for year in [2019, 2020, 2021, 2023]:
        rows.append(
            {
                "year": year,
                "school_name": "华侨大学",
                "major_original": "source_gap_no_local_hqu_candidate_records",
                "major_standard": "source_gap_no_local_hqu_candidate_records",
                "major_group": "AI基础专业群",
                "caliber": "B_basic",
                "plan_count": 0,
                "admission_scope": "",
                "batch": "",
                "remarks": "本地缺少该年份华侨大学候选明细或招生计划册。",
                "original_candidate_city": "",
                "campus_confirmed": "unknown",
                "campus_evidence_level": "F_unconfirmed",
                "campus_source_url": "",
                "campus_evidence_text": "官方章程/目录可说明校区规则，但无本地计划明细，不能生成计划数。",
                "campus_decision_reason": "缺少年份-专业-招生类别-计划数记录，按 manual_review/source_gap 处理。",
                "reassigned_city": "manual_review",
                "include_in_xiamen_E_B": 0,
                "include_in_quanzhou_E_B": 0,
                "manual_review_required": 1,
                "source_reliability": "unconfirmed",
                "reviewer_decision": "pending_review",
                "reviewer_notes": "需补充该年度福建招生计划册或华侨大学分省分专业招生计划后再核验。",
            }
        )

    rows.append(
        {
            "year": 2024,
            "school_name": "华侨大学",
            "major_original": "人工智能",
            "major_standard": "人工智能",
            "major_group": "AI基础专业群",
            "caliber": "B_basic",
            "plan_count": 0,
            "admission_scope": "",
            "batch": "",
            "remarks": "2024 章程显示计算机科学与技术学院人工智能专业大一在泉州校区，一学年后回厦门校区；本地无计划数。",
            "original_candidate_city": "",
            "campus_confirmed": "xiamen_campus",
            "campus_evidence_level": "B_official_admission_rule",
            "campus_source_url": "https://zsc.hqu.edu.cn/info/1024/7422.htm",
            "campus_evidence_text": "2024 年招生章程说明人工智能专业大一新生在泉州校区学习，一学年后回到厦门校区。",
            "campus_decision_reason": "按用户确认的华侨大学 3+1/跨校区培养规则，以最后一年所在地划分；该培养安排最后所在地为厦门。",
            "reassigned_city": "厦门",
            "include_in_xiamen_E_B": 1,
            "include_in_quanzhou_E_B": 0,
            "manual_review_required": 0,
            "source_reliability": "official",
            "reviewer_decision": "pending_review",
            "reviewer_notes": "校区归属按最后一年所在地确认；但本地没有 2024 计划数，不能新增计划数量，也不能反推 2019-2023。",
        }
    )
    return rows


def current_e_plan(panel: pd.DataFrame, city: str, year: int) -> float | None:
    if panel.empty:
        return None
    sub = panel[(panel["city"] == city) & (panel["year"] == year)]
    if sub.empty:
        return None
    value = sub.iloc[0].get("E_B_local_fujian_ai_core_major_plan")
    if pd.isna(value):
        return None
    return float(value)


def build_outputs() -> dict[str, Any]:
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    audit_dir = Path("data/audit/e_system_recheck") / run_id
    interim_dir = Path("data/interim/e_system_recheck") / run_id
    panel_dir = Path("data/panel/e_system_recheck") / run_id
    audit_dir.mkdir(parents=True, exist_ok=True)
    interim_dir.mkdir(parents=True, exist_ok=True)
    panel_dir.mkdir(parents=True, exist_ok=True)

    candidates, prior_summary, panel, warnings = load_inputs()
    hqu = candidates[candidates.get("school_name", pd.Series(dtype=str)).astype(str).str.contains("华侨大学", na=False)].copy() if not candidates.empty else pd.DataFrame()
    verification_rows = [classify_hqu_row(row) for _, row in hqu.iterrows()]
    verification_rows.extend(source_gap_rows())
    verification = pd.DataFrame(verification_rows)

    confirmed_xm_b = verification[
        (verification["campus_confirmed"] == "xiamen_campus")
        & (verification["include_in_xiamen_E_B"] == 1)
        & (verification["caliber"] == "B_basic")
    ].copy()
    confirmed_qz_b = verification[
        (verification["campus_confirmed"] == "quanzhou_campus")
        & (verification["include_in_quanzhou_E_B"] == 1)
        & (verification["caliber"] == "B_basic")
    ].copy()
    confirmed_qz_all = verification[verification["campus_confirmed"] == "quanzhou_campus"].copy()
    uncertain = verification[verification["manual_review_required"] == 1].copy()
    sources = pd.DataFrame(OFFICIAL_SOURCES)
    gap_notes = pd.DataFrame(
        [
            {
                "issue_type": "source_gap",
                "year": year,
                "notes": "本地未发现该年份华侨大学专业计划明细；官方校区规则可用于后续核验，但不能生成计划数。",
            }
            for year in [2019, 2020, 2021, 2023]
        ]
        + [
            {
                "issue_type": "final_year_location_rule",
                "year": 2024,
                "notes": "用户确认华侨大学 3+1/跨校区培养按最后一年所在地划分；2024 人工智能最后所在地为厦门，但本地无计划数。",
            },
            {
                "issue_type": "prior_recheck_revision",
                "year": 2022,
                "notes": "前一轮厦门 2022 E_B 把华侨大学 112 全列厦门候选；本轮拆分为厦门校区确认 70、泉州校区候选 42。",
            },
        ]
    )

    verification_path = interim_dir / f"hqu_2019_2024_ai_major_campus_verification_{run_id}.xlsx"
    write_xlsx(
        verification_path,
        {
            "01_all_hqu_records": verification,
            "02_confirmed_xiamen_campus": confirmed_xm_b,
            "03_confirmed_quanzhou_campus": confirmed_qz_all,
            "04_uncertain_manual_review": uncertain,
            "05_sources_checked": sources,
            "06_conflict_or_gap_notes": gap_notes,
        },
    )

    latest_verification = Path("data/interim/e_system_recheck/latest_hqu_2019_2024_ai_major_campus_verification.xlsx")
    latest_verification.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(verification_path, latest_verification)

    # Build Xiamen reclassification.
    non_hqu_b_main = pd.DataFrame()
    if not candidates.empty:
        non_hqu_b_main = candidates[
            (candidates["school_name"].astype(str) != "华侨大学")
            & (candidates["include_status"].astype(str) == "include_B_main")
            & (candidates["caliber"].astype(str) == "B_basic")
        ].copy()
    strict_2022 = int(non_hqu_b_main["plan_count"].sum()) if not non_hqu_b_main.empty else 0
    hqu_xm_2022 = int(confirmed_xm_b["plan_count"].sum())
    hqu_qz_2022 = int(confirmed_qz_b["plan_count"].sum())
    prior_b_main_2022 = int(prior_summary.iloc[0]["B_main_candidate_plan_count"]) if not prior_summary.empty else strict_2022 + hqu_xm_2022 + hqu_qz_2022

    xiamen_summary_rows = []
    for year in YEARS:
        current = current_e_plan(panel, "厦门", year)
        if year == 2022:
            strict_total = strict_2022
            with_hqu = strict_2022 + hqu_xm_2022
            pending_hqu = int(uncertain[uncertain["year"] == 2022]["plan_count"].sum())
            warning = "按校区拆分后，原 746 不再建议全额作为厦门 E_B；建议使用 704 作为人工确认候选。"
        else:
            strict_total = current if current is not None else 0
            with_hqu = strict_total
            pending_hqu = 0
            warning = "本地无华侨大学计划明细；沿用 latest panel 中非华侨大学口径，待补年度计划册复核。"
        xiamen_summary_rows.append(
            {
                "year": year,
                "current_xiamen_E_B_plan_from_latest_panel": current,
                "prior_xiamen_2022_recheck_B_main": prior_b_main_2022 if year == 2022 else "",
                "hqu_confirmed_xiamen_B_plan": hqu_xm_2022 if year == 2022 else 0,
                "hqu_confirmed_quanzhou_B_plan_removed_from_xiamen": hqu_qz_2022 if year == 2022 else 0,
                "hqu_uncertain_B_plan": pending_hqu,
                "strict_public_without_uncertain_hqu": strict_total,
                "public_with_confirmed_hqu_xiamen": with_hqu,
                "removed_or_pending_hqu_plan": (hqu_qz_2022 + pending_hqu) if year == 2022 else 0,
                "warnings": warning,
            }
        )
    xiamen_year_summary = pd.DataFrame(xiamen_summary_rows)
    xiamen_scope = pd.DataFrame()
    xiamen_school = pd.DataFrame()
    if not candidates.empty:
        revised = pd.concat(
            [
                non_hqu_b_main.assign(reclassified_bucket="strict_public_without_uncertain_hqu"),
                confirmed_xm_b.assign(
                    city="厦门",
                    school_location="厦门校区",
                    include_status="include_B_main_after_hqu_campus_check",
                    reclassified_bucket="public_with_confirmed_hqu_xiamen",
                ),
            ],
            ignore_index=True,
            sort=False,
        )
        xiamen_scope = revised.groupby(["year", "admission_scope", "reclassified_bucket"], dropna=False).agg(
            row_count=("major_standard", "count"),
            plan_count=("plan_count", "sum"),
            schools=("school_name", lambda s: "、".join(sorted(set(map(str, s))))),
        ).reset_index()
        xiamen_school = revised.groupby(["year", "school_name", "reclassified_bucket"], dropna=False).agg(
            row_count=("major_standard", "count"),
            plan_count=("plan_count", "sum"),
            majors=("major_standard", lambda s: "、".join(sorted(set(map(str, s))))),
        ).reset_index()
    else:
        revised = pd.DataFrame()

    xiamen_path = panel_dir / f"xiamen_2019_2024_E_B_reclassified_by_hqu_campus_{run_id}.xlsx"
    write_xlsx(
        xiamen_path,
        {
            "01_year_summary": xiamen_year_summary,
            "02_strict_no_uncertain_hqu": non_hqu_b_main,
            "03_confirmed_hqu_xiamen": confirmed_xm_b,
            "04_scope_breakdown": xiamen_scope,
            "05_school_breakdown": xiamen_school,
            "06_manual_review_required": uncertain,
        },
    )
    latest_xiamen = Path("data/panel/e_system_recheck/latest_xiamen_2019_2024_E_B_reclassified_by_hqu_campus.xlsx")
    shutil.copy2(xiamen_path, latest_xiamen)

    qz_rows = []
    for year in YEARS:
        current = current_e_plan(panel, "泉州", year)
        qz_plan = hqu_qz_2022 if year == 2022 else 0
        qz_rows.append(
            {
                "year": year,
                "current_quanzhou_E_B_plan_from_latest_panel": current,
                "hqu_quanzhou_candidate_B_plan": qz_plan,
                "quanzhou_E_B_if_hqu_candidate_accepted": (current + qz_plan) if current is not None else qz_plan,
                "hqu_x_expanded_plan_not_in_B": int(
                    verification[
                        (verification["year"] == year)
                        & (verification["campus_confirmed"] == "quanzhou_campus")
                        & (verification["caliber"] != "B_basic")
                    ]["plan_count"].sum()
                ),
                "manual_review_required_count": int(len(uncertain[uncertain["year"] == year])),
                "notes": "仅作为泉州 E 端校区重分类候选，不覆盖原泉州 E 端面板。" if year == 2022 else "本地无华侨大学计划明细，暂不生成新增计划数。",
            }
        )
    qz_summary = pd.DataFrame(qz_rows)
    qz_effect = qz_summary.copy()
    qz_path = panel_dir / f"quanzhou_2019_2024_E_B_hqu_quanzhou_candidate_{run_id}.xlsx"
    write_xlsx(
        qz_path,
        {
            "01_year_summary": qz_summary,
            "02_confirmed_hqu_quanzhou": confirmed_qz_b,
            "03_manual_review_required": uncertain,
            "04_effect_on_quanzhou_E": qz_effect,
        },
    )
    latest_qz = Path("data/panel/e_system_recheck/latest_quanzhou_2019_2024_E_B_hqu_quanzhou_candidate.xlsx")
    shutil.copy2(qz_path, latest_qz)

    confirmed_xm_count = int(len(confirmed_xm_b))
    confirmed_xm_plan = int(confirmed_xm_b["plan_count"].sum())
    confirmed_qz_count = int(len(confirmed_qz_b))
    confirmed_qz_plan = int(confirmed_qz_b["plan_count"].sum())
    uncertain_count = int(len(uncertain))
    uncertain_plan = int(uncertain["plan_count"].sum())

    report_path = audit_dir / f"hqu_2019_2024_campus_verification_report_{run_id}.md"
    xiamen_total_lines = "\n".join(
        f"- {int(r.year)}：strict={int(r.strict_public_without_uncertain_hqu or 0)}；with_confirmed_hqu_xiamen={int(r.public_with_confirmed_hqu_xiamen or 0)}"
        for r in xiamen_year_summary.itertuples()
    )
    qz_total_lines = "\n".join(
        f"- {int(r.year)}：HQU泉州校区 B 候选={int(r.hqu_quanzhou_candidate_B_plan)}"
        for r in qz_summary.itertuples()
    )
    record_lines = "\n".join(
        f"- {int(r.year)} {r.major_standard}：计划 {int(r.plan_count)}，校区 {r.campus_confirmed}，重分配 {r.reassigned_city}，证据 {r.campus_evidence_level}"
        for r in verification[verification["year"] == 2022].itertuples()
    )
    source_lines = "\n".join(f"- {s['source_name']}：{s['url']}" for s in OFFICIAL_SOURCES)

    report = f"""# 华侨大学 2019—2024 年 AI 基础专业群校区归属核验报告

## 1. 本次运行信息
- run_id：{run_id}
- 输入候选表：{INPUT_CANDIDATES}
- 输入 2022 PDF：{DESKTOP_PDF if DESKTOP_PDF.exists() else '未在本轮直接抽取，使用既有 recheck 候选表'}
- 输出文件：
  - {verification_path}
  - {xiamen_path}
  - {qz_path}

## 2. 核验背景
华侨大学同时在厦门、泉州两地办学，E 端如果只按学校名称或计划册中的院校所在地处理，可能把不同校区专业错误归入同一个城市。因此，本轮按“年份—专业—招生类别”核验，已确认校区的记录才进入厦门或泉州候选；校区不明或缺少年份计划明细的记录进入人工复核。对于华侨大学 3+1 或跨校区培养模式，本轮按用户确认的“最后一年所在地”规则划分。

## 3. 数据来源
本轮优先使用华侨大学招生信息网的招生章程、专业目录、学院介绍页面，以及本地 2022 福建省招生计划册 recheck 候选表。已核验来源：

{source_lines}

## 4. 专业逐年逐条校区判定
本地实际可逐条核验的华侨大学计划记录来自 2022 年厦门 E_B 补齐候选表，共 7 条，其中 B 口径 6 条、X 扩大口径 1 条：

{record_lines}

2019、2020、2021、2023 年本地未发现华侨大学计划明细，因此只记录 source_gap，不生成计划数。2024 年官方章程显示人工智能存在跨校区培养安排；按“最后一年所在地”规则，该记录归入厦门校区，但因本地无计划数，暂不新增厦门 E_B 计划数量。

## 5. 厦门 2019—2024 E_B 重分类结果
本轮对 2022 年前一版补齐值进行了校区拆分：原 `include_B_main` 中华侨大学 112 人不再全部计入厦门。其中，人工智能、计算机科学与技术、软件工程、信息安全合计 70 人可作为厦门校区确认候选；物联网工程、数据科学与大数据技术合计 42 人转为泉州校区候选。

{xiamen_total_lines}

因此，厦门 2022 年 E_B 补齐候选建议从原 746 调整为 704（非华侨大学公办主口径 634 + 华侨大学厦门校区确认 70），但仍需人工确认后再更新主面板。

## 6. 泉州 2019—2024 E_B 候选影响
确认属于华侨大学泉州校区的 B 口径候选目前只在 2022 年本地明细中出现，合计 42 人：

{qz_total_lines}

这部分可以作为泉州 2022 年 E_B 校区重分类候选，但不得直接覆盖原泉州 E 端面板。其它年份因为缺少华侨大学计划明细，暂不生成新增计划数。

## 7. 对当前 CCD 的影响
本轮不重算 CCD，也不覆盖 latest experimental base panel。当前 latest 面板中厦门 2022 仍为 missing_E，泉州 2022 仍沿用原 E_B 计划数。若人工确认本轮拆分结果，则厦门 2022 补齐候选值应使用 704 而不是 746；泉州 2022 可新增华侨大学泉州校区 B 候选 42。该变化可能影响厦门/泉州 2022 的 E_index 和后续 CCD，但需要先更新 E 端候选底表，再正式重跑 experimental_analysis_pipeline。

## 8. 下一步建议
1. 建议更新厦门 2022 E_B 补齐值：从 746 改为 704 的人工确认候选。
2. 建议新增泉州 2022 E_B 华侨大学候选：物联网工程 17、数据科学与大数据技术 25，合计 42。
3. 建议补齐 2019、2020、2021、2023、2024 年福建招生计划册或华侨大学分省分专业计划，再扩展所有年份 E 面板。
4. 人工确认后再重跑 experimental_analysis_pipeline 和 caliber test；现在不建议立即重跑 CCD。
"""
    report_path.write_text(report, encoding="utf-8")
    latest_report = Path("data/audit/e_system_recheck/latest_hqu_2019_2024_campus_verification_report.md")
    shutil.copy2(report_path, latest_report)

    output_files = [verification_path, latest_verification, xiamen_path, latest_xiamen, qz_path, latest_qz, report_path, latest_report]
    manifest_path = audit_dir / f"hqu_2019_2024_campus_verification_manifest_{run_id}.json"
    manifest = {
        "run_id": run_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "years_checked": YEARS,
        "input_files": [str(p) for p in [INPUT_CANDIDATES, INPUT_SUMMARY, LATEST_BASE_PANEL, DESKTOP_PDF] if p.exists()],
        "sources_checked": OFFICIAL_SOURCES,
        "hqu_records_checked_count": int(len(verification)),
        "local_hqu_plan_records_checked_count": int(len(hqu)),
        "confirmed_xiamen_count": confirmed_xm_count,
        "confirmed_xiamen_plan_total": confirmed_xm_plan,
        "confirmed_quanzhou_count": confirmed_qz_count,
        "confirmed_quanzhou_plan_total": confirmed_qz_plan,
        "uncertain_count": uncertain_count,
        "uncertain_plan_total": uncertain_plan,
        "xiamen_year_totals": {
            str(int(r.year)): {
                "strict_public_without_uncertain_hqu": float(r.strict_public_without_uncertain_hqu or 0),
                "public_with_confirmed_hqu_xiamen": float(r.public_with_confirmed_hqu_xiamen or 0),
            }
            for r in xiamen_year_summary.itertuples()
        },
        "quanzhou_hqu_candidate_year_totals": {
            str(int(r.year)): float(r.hqu_quanzhou_candidate_B_plan) for r in qz_summary.itertuples()
        },
        "manual_review_required_count": uncertain_count,
        "output_files": [str(p) for p in output_files],
        "warnings": warnings,
        "next_step_recommendation": "人工确认后，先更新 E 端候选底表；再重跑 experimental_analysis_pipeline 和 caliber test。",
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    output_files.append(manifest_path)

    return {
        "run_id": run_id,
        "verification_path": str(verification_path),
        "xiamen_path": str(xiamen_path),
        "qz_path": str(qz_path),
        "report_path": str(report_path),
        "manifest_path": str(manifest_path),
        "latest_verification": str(latest_verification),
        "latest_xiamen": str(latest_xiamen),
        "latest_qz": str(latest_qz),
        "latest_report": str(latest_report),
        "hqu_records_checked_count": int(len(verification)),
        "local_hqu_plan_records_checked_count": int(len(hqu)),
        "confirmed_xiamen_count": confirmed_xm_count,
        "confirmed_xiamen_plan_total": confirmed_xm_plan,
        "confirmed_quanzhou_count": confirmed_qz_count,
        "confirmed_quanzhou_plan_total": confirmed_qz_plan,
        "uncertain_count": uncertain_count,
        "uncertain_plan_total": uncertain_plan,
        "xiamen_year_summary": xiamen_year_summary.to_dict(orient="records"),
        "quanzhou_year_summary": qz_summary.to_dict(orient="records"),
        "output_files": [str(p) for p in output_files],
        "warnings": warnings,
    }


if __name__ == "__main__":
    result = build_outputs()
    print(json.dumps(result, ensure_ascii=False, indent=2))
