#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成 E 端三市 2019-2024 B 基本口径补齐版面板。

本脚本只生成 e_system_recheck 输出，不重算 CCD，不覆盖
data/panel/experimental_analysis/latest_experimental_analysis_base_panel_2019_2024.csv。
"""

from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CITIES = ["福州", "厦门", "泉州"]
YEARS = list(range(2019, 2025))

BASE_PANEL = Path("data/panel/experimental_analysis/latest_experimental_analysis_base_panel_2019_2024.csv")
HQU_VERIFICATION = Path("data/interim/e_system_recheck/latest_hqu_2019_2024_ai_major_campus_verification.xlsx")
XIAMEN_RECLASSIFIED = Path("data/panel/e_system_recheck/latest_xiamen_2019_2024_E_B_reclassified_by_hqu_campus.xlsx")
QUANZHOU_HQU = Path("data/panel/e_system_recheck/latest_quanzhou_2019_2024_E_B_hqu_quanzhou_candidate.xlsx")
XIAMEN_2022_RECHECK = Path("data/panel/e_system_recheck/latest_xiamen_2022_E_B_city_year_summary.xlsx")
HQU_LATEST_REPORT = Path("data/audit/e_system_recheck/latest_hqu_2019_2024_campus_verification_report.md")


def clean(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def latest_hqu_run_id() -> str:
    if not HQU_LATEST_REPORT.exists():
        return "unknown_hqu_latest"
    text = HQU_LATEST_REPORT.read_text(encoding="utf-8")
    match = re.search(r"run_id[：:]\s*([0-9]{8}_[0-9]{6})", text)
    return match.group(1) if match else "unknown_hqu_latest"


def style_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col[:80]:
                max_len = max(max_len, min(len(clean(cell.value)), 70))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 42))
    wb.save(path)


def write_xlsx(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    style_workbook(path)


def load_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, list[str]]:
    warnings: list[str] = []
    if not BASE_PANEL.exists():
        raise FileNotFoundError(f"缺少原 experimental base panel: {BASE_PANEL}")
    base = pd.read_csv(BASE_PANEL)

    if XIAMEN_RECLASSIFIED.exists():
        xiamen = pd.read_excel(XIAMEN_RECLASSIFIED, sheet_name="01_year_summary")
    else:
        xiamen = pd.DataFrame()
        warnings.append(f"缺少厦门 HQU 重分类汇总: {XIAMEN_RECLASSIFIED}")

    if QUANZHOU_HQU.exists():
        quanzhou = pd.read_excel(QUANZHOU_HQU, sheet_name="01_year_summary")
    else:
        quanzhou = pd.DataFrame()
        warnings.append(f"缺少泉州 HQU 候选汇总: {QUANZHOU_HQU}")

    if HQU_VERIFICATION.exists():
        hqu = pd.read_excel(HQU_VERIFICATION, sheet_name="01_all_hqu_records")
    else:
        hqu = pd.DataFrame()
        warnings.append(f"缺少 HQU 校区核验明细: {HQU_VERIFICATION}")

    if not XIAMEN_2022_RECHECK.exists():
        warnings.append(f"缺少厦门 2022 E_B 原补齐汇总: {XIAMEN_2022_RECHECK}")

    warnings.append("本轮未读取 archive/legacy_outputs，未覆盖 latest experimental base panel，未重算 CCD。")
    return base, xiamen, quanzhou, hqu, warnings


def get_xiamen_2022_value(xiamen: pd.DataFrame) -> int:
    if xiamen.empty:
        return 704
    row = xiamen[xiamen["year"].astype(int).eq(2022)]
    if row.empty:
        return 704
    value = row.iloc[0].get("public_with_confirmed_hqu_xiamen")
    return int(value) if not pd.isna(value) else 704


def get_quanzhou_2022_value(quanzhou: pd.DataFrame, original: float) -> int:
    if quanzhou.empty:
        return int(original)
    row = quanzhou[quanzhou["year"].astype(int).eq(2022)]
    if row.empty:
        return int(original)
    value = row.iloc[0].get("quanzhou_E_B_if_hqu_candidate_accepted")
    return int(value) if not pd.isna(value) else int(original)


def build_panel() -> dict[str, Any]:
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    source_run_id = latest_hqu_run_id()
    panel_dir = Path("data/panel/e_system_recheck") / run_id
    audit_dir = Path("data/audit/e_system_recheck") / run_id
    panel_dir.mkdir(parents=True, exist_ok=True)
    audit_dir.mkdir(parents=True, exist_ok=True)

    base, xiamen, quanzhou, hqu, warnings = load_inputs()
    sub = base[base["city"].isin(CITIES) & base["year"].isin(YEARS)].copy()
    sub = sub.sort_values(["city", "year"]).reset_index(drop=True)

    original_col = "E_B_local_fujian_ai_core_major_plan"
    xiamen_2022 = get_xiamen_2022_value(xiamen)
    qz_original_2022 = sub[(sub["city"] == "泉州") & (sub["year"] == 2022)].iloc[0][original_col]
    quanzhou_2022 = get_quanzhou_2022_value(quanzhou, float(qz_original_2022))

    completed_rows: list[dict[str, Any]] = []
    for _, row in sub.iterrows():
        city = row["city"]
        year = int(row["year"])
        original = row.get(original_col)
        corrected = original
        correction_reason = "unchanged_from_latest_experimental_base_panel"
        row_source_run_id = "latest_experimental_analysis_base_panel"
        corrected_status = row.get("data_status")
        include_hqu_campus_reclassification = 0

        if city == "厦门" and year == 2022:
            corrected = xiamen_2022
            correction_reason = (
                "fill_missing_xiamen_2022_E_B_with_hqu_campus_reclassified_value: "
                "strict_public_without_uncertain_hqu=634 + confirmed_hqu_xiamen=70"
            )
            row_source_run_id = source_run_id
            corrected_status = "calculated"
            include_hqu_campus_reclassification = 1
        elif city == "泉州" and year == 2022:
            corrected = quanzhou_2022
            correction_reason = (
                "hqu_campus_reclassification_add_confirmed_quanzhou_B_candidate: "
                "original_quanzhou_E_B=1324 + confirmed_hqu_quanzhou=42"
            )
            row_source_run_id = source_run_id
            corrected_status = "calculated"
            include_hqu_campus_reclassification = 1

        original_missing = pd.isna(original)
        corrected_missing = pd.isna(corrected)
        completed_rows.append(
            {
                "city": city,
                "year": year,
                "original_E_B": original,
                "corrected_E_B": corrected,
                "correction_delta": None if original_missing else float(corrected) - float(original),
                "original_data_status": row.get("data_status"),
                "corrected_data_status": corrected_status,
                "original_E_B_missing": int(original_missing),
                "corrected_E_B_missing": int(corrected_missing),
                "correction_reason": correction_reason,
                "source_run_id": row_source_run_id,
                "hqu_campus_reclassification_used": include_hqu_campus_reclassification,
                "original_E_index": row.get("E_index"),
                "original_school_list": row.get("学校列表"),
                "notes": "补齐版仅更新 E_B 主值；E_index/CCD 需通过后续 experimental_analysis_pipeline 重算。" if include_hqu_campus_reclassification else "",
            }
        )
    completed = pd.DataFrame(completed_rows)

    missing_check = completed.assign(
        e_b_complete_after=completed["corrected_E_B_missing"].eq(0).astype(int),
        issue_after=completed["corrected_E_B_missing"].map({0: "", 1: "missing_corrected_E_B"}),
    )[
        [
            "city",
            "year",
            "original_E_B",
            "corrected_E_B",
            "original_E_B_missing",
            "corrected_E_B_missing",
            "e_b_complete_after",
            "original_data_status",
            "corrected_data_status",
            "correction_reason",
            "source_run_id",
            "issue_after",
        ]
    ]

    city_year_summary = completed.groupby("city", dropna=False).agg(
        year_count=("year", "count"),
        corrected_E_B_non_missing_count=("corrected_E_B_missing", lambda s: int((s == 0).sum())),
        corrected_E_B_missing_count=("corrected_E_B_missing", "sum"),
        min_corrected_E_B=("corrected_E_B", "min"),
        max_corrected_E_B=("corrected_E_B", "max"),
    ).reset_index()
    city_year_summary["is_complete_2019_2024"] = (
        (city_year_summary["year_count"] == 6)
        & (city_year_summary["corrected_E_B_missing_count"] == 0)
    ).astype(int)

    hqu_summary = pd.DataFrame(
        [
            {
                "item": "xiamen_2022_original_missing_filled",
                "value": xiamen_2022,
                "notes": "厦门 2022 corrected_E_B 主值设置为 704。",
            },
            {
                "item": "quanzhou_2022_hqu_confirmed_added",
                "value": quanzhou_2022,
                "notes": "泉州 2022 corrected_E_B = 1324 + 华侨大学泉州校区确认 B 候选 42。",
            },
            {
                "item": "hqu_uncertain_records_included_in_main",
                "value": 0,
                "notes": "校区不明或 source_gap 记录不进入主值。",
            },
        ]
    )

    source_files = pd.DataFrame(
        [
            {"source_name": "latest_experimental_base_panel", "path": str(BASE_PANEL), "role": "original_E_B source"},
            {"source_name": "latest_hqu_campus_verification", "path": str(HQU_VERIFICATION), "role": "HQU campus mapping"},
            {"source_name": "latest_xiamen_reclassified_by_hqu", "path": str(XIAMEN_RECLASSIFIED), "role": "Xiamen 2022 corrected_E_B=704"},
            {"source_name": "latest_quanzhou_hqu_candidate", "path": str(QUANZHOU_HQU), "role": "Quanzhou 2022 corrected_E_B=1366"},
            {"source_name": "latest_xiamen_2022_E_B_recheck", "path": str(XIAMEN_2022_RECHECK), "role": "prior Xiamen 2022 candidate value=746"},
        ]
    )

    out_path = panel_dir / f"E_B_three_city_2019_2024_completed_panel_{run_id}.xlsx"
    write_xlsx(
        out_path,
        {
            "01_completed_panel": completed,
            "02_E_B_missing_check": missing_check,
            "03_city_completion_summary": city_year_summary,
            "04_hqu_correction_summary": hqu_summary,
            "05_source_files": source_files,
            "06_hqu_records_reference": hqu,
        },
    )
    latest_path = Path("data/panel/e_system_recheck/latest_E_B_three_city_2019_2024_completed_panel.xlsx")
    shutil.copy2(out_path, latest_path)

    all_complete = bool((completed["corrected_E_B_missing"] == 0).all() and len(completed) == 18)
    missing_after = missing_check[missing_check["corrected_E_B_missing"].eq(1)]
    corrected_rows = completed[completed["hqu_campus_reclassification_used"].eq(1)]

    report_path = audit_dir / f"E_B_three_city_completion_report_{run_id}.md"
    summary_lines = "\n".join(
        f"- {r.city}：{int(r.corrected_E_B_non_missing_count)}/6 年有 corrected_E_B，缺失 {int(r.corrected_E_B_missing_count)} 年。"
        for r in city_year_summary.itertuples()
    )
    corrected_lines = "\n".join(
        f"- {r.city} {int(r.year)}：original_E_B={'' if pd.isna(r.original_E_B) else int(r.original_E_B)}，corrected_E_B={int(r.corrected_E_B)}，原因：{r.correction_reason}"
        for r in corrected_rows.itertuples()
    )
    report = f"""# E 端三市 2019—2024 B 基本口径补齐版面板报告

## 1. 本次运行信息
- run_id：{run_id}
- source_run_id：{source_run_id}
- 输出面板：{out_path}
- latest 面板：{latest_path}
- 原 experimental base panel 未覆盖：{BASE_PANEL}

## 2. 数据使用说明
本轮只基于 latest experimental base panel、华侨大学校区核验 latest、厦门/泉州 HQU 重分类 latest 生成 E 端补齐版面板。不重算 CCD，不覆盖原 latest experimental base panel。

## 3. 补齐与重分类规则
- 厦门 2022 `corrected_E_B` 设为 704，即非华侨大学公办主口径 634 + 华侨大学厦门校区确认 B 口径 70。
- 泉州 2022 `corrected_E_B` 设为 1366，即原泉州 1324 + 华侨大学泉州校区确认 B 候选 42。
- 校区不明、source_gap、无计划数记录不进入主值。
- 华侨大学 3+1/跨校区培养按最后一年所在地划分；但无年度计划数时不新增计划数量。

## 4. 三市 E_B 完整性
{summary_lines}

结论：E 端三市小样本 2019—2024 B 基本口径在补齐版面板中{'已经完整' if all_complete else '仍未完整'}。

## 5. 厦门 2022 状态
厦门 2022 已由原 experimental base panel 中的 `missing_E` / `E_B` 缺失，修正为 `corrected_data_status=calculated`，`corrected_E_B=704`。这只是 E 端补齐版面板状态，不代表已重算 E_index 或 CCD。

## 6. 华侨大学校区重分类影响
{corrected_lines}

华侨大学原先在厦门 2022 补齐候选中合计 112 人，本轮按确认校区拆分：厦门校区 B 口径 70 人进入厦门主值；泉州校区 B 候选 42 人进入泉州修正值。校区不明记录不进入主值。

## 7. 是否还存在 E_B 缺失
补齐后 corrected_E_B 缺失数量为 {int(completed['corrected_E_B_missing'].sum())}。{'不存在 E_B 缺失。' if missing_after.empty else '仍存在 E_B 缺失，详见 02_E_B_missing_check。'}

## 8. 下一步建议
建议下一步重新跑 `experimental_analysis_pipeline`，用本轮 `corrected_E_B` 更新 E 端主输入后重新计算 E_index、D 和 CCD，并随后重跑 caliber test。当前文件只是 E 端 B 基本口径补齐面板，不应直接替代 CCD 结果表。
"""
    report_path.write_text(report, encoding="utf-8")

    return {
        "run_id": run_id,
        "source_run_id": source_run_id,
        "output_panel": str(out_path),
        "latest_panel": str(latest_path),
        "report": str(report_path),
        "row_count": int(len(completed)),
        "all_complete": all_complete,
        "missing_after_count": int(completed["corrected_E_B_missing"].sum()),
        "xiamen_2022_corrected_E_B": xiamen_2022,
        "quanzhou_2022_corrected_E_B": quanzhou_2022,
        "warnings": warnings,
    }


if __name__ == "__main__":
    result = build_panel()
    import json

    print(json.dumps(result, ensure_ascii=False, indent=2))
